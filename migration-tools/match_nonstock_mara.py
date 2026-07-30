"""Cocokkan material Non-Stock (AppSheet) yang belum ada di katalog WARNOTO
dengan usulan kode+deskripsi katalog MARA, untuk review manual admin.

Reusable untuk UPT mana pun via --upt. Tidak menulis apa pun ke Supabase.
"""
import argparse
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

DEFAULT_SOURCE = r"D:\CLAUDE\WARNOTO data\Appsheet\Filter Material Non Stock.xlsx"
DEFAULT_MARA = r"D:\CLAUDE\WARNOTO data\tester\Katalog MARA (01-2026).xlsx"
DEFAULT_EXISTING = (
    r"D:\CLAUDE\WARNOTO CODE\warnoto-project\outputs\warnoto-nonstock-review"
    r"\_existing_katalog_live_2026-07-30.json"
)
PHOTO_ROOT = r"D:\CLAUDE\WARNOTO data\Appsheet\_extracted\data"

SHEET_NEW = "Usulan Pencocokan MARA"
SHEET_REVIEW = "Qty 0 atau Kosong (Review)"


def normalize_code(value):
    """Kode katalog -> string tanpa desimal. Return (kode, valid)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "", False
    try:
        return str(int(float(value))), True
    except (ValueError, TypeError):
        text = str(value).strip()
        return text, False


def tokenize(text):
    if not text:
        return set()
    return {t for t in re.split(r"[^A-Z0-9]+", str(text).upper()) if t}


def build_mara_index(mara_df):
    """Precompute: exact-code lookup + token set per baris MARA."""
    by_code = {}
    entries = []  # (kode, deskripsi, token_set)
    for _, row in mara_df.iterrows():
        kode, valid = normalize_code(row["Material"])
        desc = "" if pd.isna(row["Material Description"]) else str(row["Material Description"]).strip()
        if valid and kode not in by_code:
            by_code[kode] = desc
        entries.append((kode, desc, tokenize(desc)))
    return by_code, entries


def find_keyword_candidates(nama_material, entries):
    """Top 3 kandidat MARA berdasarkan overlap token, ranked by (overlap desc, len token asc)."""
    query_tokens = tokenize(nama_material)
    if not query_tokens:
        return []
    scored = []
    for kode, desc, tokens in entries:
        overlap = len(query_tokens & tokens)
        if overlap == 0:
            continue
        scored.append((overlap, len(tokens), kode, desc))
    scored.sort(key=lambda x: (-x[0], x[1]))
    top = scored[:3]
    coverage = top[0][0] / max(1, len(query_tokens)) if top else 0
    return top, coverage


def match_row(nama_material, katalog_norm, code_valid, mara_by_code, mara_entries):
    if code_valid and katalog_norm in mara_by_code:
        return {
            "method": "Kode Persis",
            "kode": katalog_norm,
            "deskripsi": mara_by_code[katalog_norm],
            "skor": "COCOK KODE",
            "lain": "",
        }

    result = find_keyword_candidates(nama_material, mara_entries)
    if not result or not result[0]:
        return {"method": "Kata Kunci Deskripsi", "kode": "", "deskripsi": "", "skor": "TANPA KANDIDAT", "lain": ""}

    top, coverage = result
    best_overlap, _, kode1, desc1 = top[0]
    if best_overlap == 0:
        return {"method": "Kata Kunci Deskripsi", "kode": "", "deskripsi": "", "skor": "TANPA KANDIDAT", "lain": ""}

    skor = "KUAT" if coverage >= 0.75 else "LEMAH"
    lain_parts = []
    for overlap, _, kode, desc in top[1:3]:
        lain_parts.append(f"{kode} | {desc} ({overlap} kata cocok)")
    return {
        "method": "Kata Kunci Deskripsi",
        "kode": kode1,
        "deskripsi": desc1,
        "skor": skor,
        "lain": "; ".join(lain_parts),
    }


def exclude_existing(df, existing_codes):
    """Tambah kolom kode-normalisasi + pisahkan baris yang kodenya sudah ada di katalog WARNOTO."""
    df = df.copy()
    codes = df["Katalog"].apply(normalize_code)
    df["_kode_norm"] = [c[0] for c in codes]
    df["_kode_valid"] = [c[1] for c in codes]
    df["_is_existing"] = df["_kode_norm"].isin(existing_codes)
    excluded = df[df["_is_existing"]]
    remaining = df[~df["_is_existing"]].copy()
    return excluded, remaining


def build_photo_index(root_dir):
    """Scan semua subfolder foto backup AppSheet sekali di awal -> {nama_file: path_absolut}.
    Duplikat nama file antar folder: yang ketemu pertama menang (isinya identik, sudah dicek user)."""
    index = {}
    root = Path(root_dir)
    for pattern in ("*/List Material_Images/*", "*/*Material_Images/*"):
        for path in root.glob(pattern):
            if path.is_file() and path.name not in index:
                index[path.name] = str(path.resolve())
    return index


def resolve_photo_link(rel_path, photo_index):
    """rel_path kosong/NaN -> None. Kalau ada, cari nama file (setelah '/' terakhir) di photo_index."""
    if rel_path is None or (isinstance(rel_path, float) and pd.isna(rel_path)):
        return None
    filename = str(rel_path).strip().split("/")[-1]
    if not filename:
        return None
    return photo_index.get(filename)


def build_rows(remaining, mara_by_code, mara_entries, photo_index=None):
    """Jalankan match_row() per baris, hasilkan out_rows + counts (dipakai kedua kelompok).
    photo_index diberikan -> tambah 2 kolom link foto (hanya dipakai untuk sheet Usulan Pencocokan MARA)."""
    out_rows = []
    counts = {"Kode Persis": 0, "KUAT": 0, "LEMAH": 0, "TANPA KANDIDAT": 0}
    for _, row in remaining.iterrows():
        m = match_row(row["Nama Material"], row["_kode_norm"], row["_kode_valid"], mara_by_code, mara_entries)
        if m["method"] == "Kode Persis":
            counts["Kode Persis"] += 1
        else:
            counts[m["skor"]] += 1
        out_row = {
            "idMaterial": row.get("idMaterial"),
            "Milik UPT": row.get("Milik UPT"),
            "Katalog Asli (AppSheet)": row.get("Katalog"),
            "Kode Katalog Usulan (MARA)": m["kode"],
            "Nama Material": row.get("Nama Material"),
            "Nama Usulan (MARA)": m["deskripsi"],
        }
        if photo_index is not None:
            out_row["Link Foto Material"] = resolve_photo_link(row.get("Foto Material"), photo_index)
            out_row["Link Foto Material Tambahan"] = resolve_photo_link(row.get("Foto Material Tambahan"), photo_index)
        out_row.update({
            "Satuan": row.get("Satuan"),
            "Jumlah Stok": row.get("Jumlah Stok"),
            "Status": row.get("Status"),
            "Metode Pencocokan": m["method"],
            "Skor Kecocokan": m["skor"],
            "Kandidat MARA Lain (top 2-3)": m["lain"],
            "Keputusan Admin (isi manual)": "",
            "Catatan (isi manual)": "",
        })
        out_rows.append(out_row)
    return out_rows, counts


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--mara", default=DEFAULT_MARA)
    parser.add_argument("--existing-katalog", default=DEFAULT_EXISTING)
    parser.add_argument("--upt", default="UPT Surabaya")
    parser.add_argument("--output", default=None)
    parser.add_argument("--photo-root", default=PHOTO_ROOT)
    args = parser.parse_args()

    upt_slug = args.upt.upper().replace(" ", "_")
    if upt_slug.startswith("UPT_"):
        upt_slug = upt_slug[len("UPT_"):]
    today = datetime.now().strftime("%Y%m%d")
    output_path = Path(args.output) if args.output else Path(
        f"outputs/warnoto-nonstock-review/USULAN_PENCOCOKAN_MARA_NONSTOCK_{upt_slug}_{today}.xlsx"
    )

    source_df = pd.read_excel(args.source, sheet_name="listMaterial", dtype=object)
    mara_df = pd.read_excel(args.mara, sheet_name="Sheet1", dtype=object)
    existing = pd.read_json(args.existing_katalog)
    existing_codes = {normalize_code(v)[0] for v in existing["katalog"]}

    mara_by_code, mara_entries = build_mara_index(mara_df)
    photo_index = build_photo_index(args.photo_root)

    upt_rows = source_df[source_df["Milik UPT"] == args.upt].copy()
    qty_upt = pd.to_numeric(upt_rows["Jumlah Stok"], errors="coerce")
    filtered = upt_rows[qty_upt > 0].copy()
    filtered_zero_or_empty = upt_rows[~(qty_upt > 0)].copy()
    total_filter = len(filtered)
    total_filter_review = len(filtered_zero_or_empty)

    excluded, remaining = exclude_existing(filtered, existing_codes)
    total_exclude = len(excluded)
    total_final = len(remaining)

    excluded_review, remaining_review = exclude_existing(filtered_zero_or_empty, existing_codes)
    total_exclude_review = len(excluded_review)
    total_final_review = len(remaining_review)

    out_rows, counts = build_rows(remaining, mara_by_code, mara_entries, photo_index=photo_index)
    out_df = pd.DataFrame(out_rows)

    out_rows_review, counts_review = build_rows(remaining_review, mara_by_code, mara_entries)
    out_df_review = pd.DataFrame(out_rows_review)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(Path(args.source).read_bytes())

    wb = load_workbook(output_path)
    if SHEET_NEW in wb.sheetnames:
        del wb[SHEET_NEW]
    if SHEET_REVIEW in wb.sheetnames:
        del wb[SHEET_REVIEW]
    listmaterial_idx = wb.sheetnames.index("listMaterial")
    ws = wb.create_sheet(SHEET_NEW, index=listmaterial_idx + 1)

    readme = (
        f"Generate: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"Baris awal (filter {args.upt} + Jumlah Stok > 0): {total_filter} | "
        f"Exclude (sudah ada di katalog WARNOTO): {total_exclude} | "
        f"Baris final di sheet ini: {total_final} | "
        "Katalog WARNOTO yang dipakai untuk exclude: snapshot live per 2026-07-30."
    )
    ws.append([readme])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_df.columns))
    ws.append([])

    for r in dataframe_to_rows(out_df, index=False, header=True):
        ws.append(r)

    for col_idx, col_name in enumerate(out_df.columns, start=1):
        width = max(12, min(50, len(str(col_name)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header di baris 3 (baris 1 = readme, baris 2 = kosong), data mulai baris 4.
    photo_cols = [c for c in ("Link Foto Material", "Link Foto Material Tambahan") if c in out_df.columns]
    link_font = Font(color="0563C1", underline="single")
    for col_name in photo_cols:
        col_idx = out_df.columns.get_loc(col_name) + 1
        for row_offset, path in enumerate(out_df[col_name]):
            if not path or (isinstance(path, float) and pd.isna(path)):
                continue
            cell = ws.cell(row=4 + row_offset, column=col_idx)
            cell.value = Path(path).name
            cell.hyperlink = "file:///" + path.replace("\\", "/")
            cell.font = link_font

    ws_review = wb.create_sheet(SHEET_REVIEW, index=wb.sheetnames.index(SHEET_NEW) + 1)

    readme_review = (
        f"Generate: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"Baris awal (filter {args.upt}, Jumlah Stok kosong/0/negatif): {total_filter_review} | "
        f"Exclude (sudah ada di katalog WARNOTO): {total_exclude_review} | "
        f"Baris final di sheet ini: {total_final_review} | "
        "Baris ini TIDAK memiliki Jumlah Stok > 0 (kosong/nol/negatif) - kemungkinan barang sudah tidak ada "
        "fisiknya, ATAU data AppSheet lama belum lengkap. Perlu ditinjau manual, bukan otomatis diusulkan "
        "sebagai material baru."
    )
    ws_review.append([readme_review])
    ws_review.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_df_review.columns))
    ws_review.append([])

    for r in dataframe_to_rows(out_df_review, index=False, header=True):
        ws_review.append(r)

    for col_idx, col_name in enumerate(out_df_review.columns, start=1):
        width = max(12, min(50, len(str(col_name)) + 2))
        ws_review.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)

    print(f"Output: {output_path}")
    print(f"Total filter (UPT={args.upt}, Jumlah Stok>0): {total_filter}")
    print(f"Exclude (sudah ada di katalog): {total_exclude}")
    print(f"Final di sheet baru: {total_final}")
    print(f"Breakdown final: Kode Persis={counts['Kode Persis']}, KUAT={counts['KUAT']}, "
          f"LEMAH={counts['LEMAH']}, TANPA KANDIDAT={counts['TANPA KANDIDAT']}")

    # exact-code matches among ALL 135 (before exclude), for architect sanity check
    all_codes = filtered["Katalog"].apply(normalize_code)
    all_codes_valid = [c[0] for c in all_codes if c[1]]
    exact_all = sum(1 for c in all_codes_valid if c in mara_by_code)
    print(f"Exact code match ke MARA di antara SEMUA {total_filter} baris (sebelum exclude): {exact_all}")

    print(f"[Review qty 0/kosong] Total filter (UPT={args.upt}, Jumlah Stok<=0/NaN): {total_filter_review}")
    print(f"[Review qty 0/kosong] Exclude (sudah ada di katalog): {total_exclude_review}")
    print(f"[Review qty 0/kosong] Final di sheet baru: {total_final_review}")
    print(f"[Review qty 0/kosong] Breakdown final: Kode Persis={counts_review['Kode Persis']}, "
          f"KUAT={counts_review['KUAT']}, LEMAH={counts_review['LEMAH']}, "
          f"TANPA KANDIDAT={counts_review['TANPA KANDIDAT']}")


def demo():
    """Self-check kecil untuk logic inti tanpa file eksternal."""
    assert normalize_code(1060154.0) == ("1060154", True)
    assert normalize_code("abc")[1] is False
    assert tokenize("CB ACC;CLOSING TRIP COIL 110VDC") == {"CB", "ACC", "CLOSING", "TRIP", "COIL", "110VDC"}

    mara_by_code = {"123": "FOO;BAR BAZ"}
    mara_entries = [("123", "FOO;BAR BAZ", tokenize("FOO;BAR BAZ")), ("456", "FOO;QUX", tokenize("FOO;QUX"))]

    exact = match_row("apa saja", "123", True, mara_by_code, mara_entries)
    assert exact["method"] == "Kode Persis" and exact["skor"] == "COCOK KODE"

    kw = match_row("FOO BAR BAZ", "999", True, mara_by_code, mara_entries)
    assert kw["method"] == "Kata Kunci Deskripsi" and kw["skor"] == "KUAT" and kw["kode"] == "123"

    none_match = match_row("ZZZ YYY", "999", True, mara_by_code, mara_entries)
    assert none_match["skor"] == "TANPA KANDIDAT"
    print("demo() OK")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--demo":
        demo()
    else:
        main()
